# tableau_admissions_report/report_processor/excel_formatter.py
import logging
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

from config import settings # Import formatting constants

logger = logging.getLogger(__name__)

def _auto_adjust_column_widths(ws: Worksheet, padding: int = 5, max_width: int = 100):
    """
    Adjusts column widths based on the maximum content length in each column.
    Skips columns that are part of a merged cell range for width calculation
    based on individual cells within them, as the merged cell itself dictates width.
    """
    logger.debug(f"Adjusting column widths for sheet: {ws.title}")
    # Get set of all cells that are part of a merge
    merged_cells_in_sheet = set()
    for merged_range_obj in ws.merged_cells.ranges:
        for row, col in merged_range_obj.cells:
            merged_cells_in_sheet.add(ws.cell(row=row, column=col).coordinate)

    for col_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col_cells[0].column)

        is_column_part_of_merge = False
        # Check if the first cell of this column is part of any merged cell range
        # This is a simplification; a column might be partially merged.
        # A more robust check would iterate all cells in the column.
        if col_cells[0].coordinate in merged_cells_in_sheet:
             # If the column starts with a merged cell, let the merged cell define its width,
             # or apply a default/fixed width if necessary.
             # For now, we might skip auto-sizing or apply a moderate default.
             # Let's try to still calculate based on non-merged cells in this column.
             pass


        for cell in col_cells:
            # Only consider cells not involved in merges for individual length calculation
            if cell.coordinate in merged_cells_in_sheet:
                continue
            try:
                if cell.value:
                    cell_value_str = str(cell.value)
                    if len(cell_value_str) > max_length:
                        max_length = len(cell_value_str)
            except: # pylint: disable=bare-except
                pass # Ignore errors from cell value conversion
        
        # If max_length is still 0 (e.g., empty column or all merged), use a default.
        # Otherwise, add padding. Cap at max_width.
        adjusted_width = (max_length + padding) if max_length > 0 else 15 # Default width for empty/merged columns
        ws.column_dimensions[column_letter].width = min(adjusted_width, max_width)
    logger.debug(f"Column widths adjusted for sheet: {ws.title}")


def _apply_styles_to_progress_report_sheet(ws: Worksheet):
    """
    Applies specific formatting to the 'Progress Report' sheet.
    Includes merging cells, applying fonts, fills, borders, and alignment.
    """
    logger.info(f"Applying detailed styles to 'Progress Report' sheet: {ws.title}")

    # --- Freeze top row (header) ---
    ws.freeze_panes = 'A2' # Freezes row 1
    logger.debug(f"Froze top row for sheet: {ws.title}")

    if ws.max_row <= 1: # Empty or header-only sheet
        logger.warning(f"Sheet '{ws.title}' is empty or has only headers. Applying basic header style and auto-width.")
        if ws.max_row == 1: # Style header if it exists
             for cell in ws[1]:
                cell.font = Font(bold=settings.EXCEL_FONT_BOLD)
                cell.alignment = Alignment(**settings.EXCEL_ALIGNMENT_CENTER)
        _auto_adjust_column_widths(ws)
        return

    # --- Define Styles from Config (or directly) ---
    font_bold = Font(bold=settings.EXCEL_FONT_BOLD)
    align_center_center = Alignment(**settings.EXCEL_ALIGNMENT_CENTER)
    
    fill_alt_row = PatternFill(**settings.EXCEL_FILL_ALT_ROW)
    fill_term_block_alt = PatternFill(**settings.EXCEL_FILL_ALT_ROW)
    fill_total_row = PatternFill(**settings.EXCEL_FILL_TOTAL_ROW)
    fill_grand_total_row = PatternFill(**settings.EXCEL_FILL_GRAND_TOTAL_ROW)

    side_thin_white = Side(**settings.EXCEL_BORDER_SIDE_THIN_WHITE)
    side_medium_black = Side(**settings.EXCEL_BORDER_SIDE_MEDIUM_BLACK)

    border_thin_white_all_sides = Border(left=side_thin_white, right=side_thin_white, top=side_thin_white, bottom=side_thin_white)

    # --- Column Indices (1-based) ---
    # These should align with settings.PROGRESS_REPORT_FINAL_COLUMN_ORDER
    try:
        term_col_idx = settings.PROGRESS_REPORT_FINAL_COLUMN_ORDER.index("Application Term") + 1
        program_col_idx = settings.PROGRESS_REPORT_FINAL_COLUMN_ORDER.index("Program") + 1
        degree_col_idx = settings.PROGRESS_REPORT_FINAL_COLUMN_ORDER.index("DEGREE") + 1
        first_data_col_idx = degree_col_idx + 1 # Assuming data columns start after DEGREE
    except ValueError:
        logger.error("One of the key columns for styling (Application Term, Program, DEGREE) is not in PROGRESS_REPORT_FINAL_COLUMN_ORDER. Styling may be incorrect.")
        # Fallback or raise error
        _auto_adjust_column_widths(ws) # At least adjust widths
        return


    # --- 1. Merge 'Application Term' cells ---
    logger.debug("Merging 'Application Term' cells...")
    current_term_value_being_grouped = None
    merge_start_row_for_current_term = 2 # Header is row 1

    # Iterate up to ws.max_row (inclusive). The +1 in range end is to process/flush the last group.
    for row_idx_iterator in range(2, ws.max_row + 2): 
        # Get the value of the cell in the "Application Term" column for the current row being inspected.
        # If past the actual last row, treat as None to trigger merge for the last group.
        value_in_current_cell = ws.cell(row=row_idx_iterator, column=term_col_idx).value if row_idx_iterator <= ws.max_row else None
        
        # Check if the value in the current cell is the "Grand Total" label.
        is_current_cell_grand_total_label = (value_in_current_cell == "Grand Total")

        # --- Condition to finalize and perform merge for the 'current_term_value_being_grouped' ---
        # This block executes when:
        #   a) The term value changes (value_in_current_cell != current_term_value_being_grouped)
        #   b) OR we are one step past the last actual data row (row_idx_iterator == ws.max_row + 1),
        #      which serves to flush out the merge for the very last group of terms.
        # AND
        #   c) There is a term currently being grouped (current_term_value_being_grouped is not None).
        if (value_in_current_cell != current_term_value_being_grouped or row_idx_iterator == ws.max_row + 1) and \
           current_term_value_being_grouped is not None:
            
            # We only perform this group-merge if 'current_term_value_being_grouped' was an actual academic term,
            # NOT if it was "Grand Total" itself (which has separate label merging logic).
            if current_term_value_being_grouped != "Grand Total": 
                # Determine the end row for the merge operation.
                # If the term changed at row_idx_iterator, the previous term's group ended at row_idx_iterator - 1.
                # If we are at the flush step (row_idx_iterator == ws.max_row + 1), the group ended at ws.max_row.
                end_row_for_this_merge = (row_idx_iterator - 1) if row_idx_iterator <= ws.max_row else ws.max_row
                
                # Ensure the calculated range is valid (end row is not before start row).
                if end_row_for_this_merge >= merge_start_row_for_current_term:
                    try:
                        ws.merge_cells(start_row=merge_start_row_for_current_term, 
                                       end_row=end_row_for_this_merge, 
                                       start_column=term_col_idx, 
                                       end_column=term_col_idx)
                        # Apply alignment to the top cell of the newly merged range.
                        ws.cell(row=merge_start_row_for_current_term, column=term_col_idx).alignment = align_center_center
                        logger.debug(f"Merged 'Application Term' for '{current_term_value_being_grouped}' from row {merge_start_row_for_current_term} to {end_row_for_this_merge}")
                    except Exception as e:
                        logger.warning(f"Could not merge 'Application Term' cells for '{current_term_value_being_grouped}': {e}")
        
        # --- Logic to handle transitions and stop conditions ---

        # If the current cell's value is "Grand Total":
        # The merge for the academic term *preceding* "Grand Total" would have just been processed by the block above.
        # Now, we stop further attempts to group-merge "Application Term" cells.
        if is_current_cell_grand_total_label:
            logger.debug(f"Encountered 'Grand Total' label at row {row_idx_iterator}. Stopping 'Application Term' group merging.")
            break # Exit the loop for merging Application Term groups.
            
        # If the iterator has moved past all actual data rows in the worksheet, stop.
        if row_idx_iterator > ws.max_row:
            break

        # If the value in the current cell is different from the term we were grouping,
        # it means a new term (or the first term) has started. Update tracking variables.
        if value_in_current_cell != current_term_value_being_grouped:
            current_term_value_being_grouped = value_in_current_cell
            merge_start_row_for_current_term = row_idx_iterator
            
    logger.debug("'Application Term' cell merging complete.")

    # --- Apply Alternating Fill to Merged "Application Term" Blocks ---
    logger.debug("Applying alternating fill to 'Application Term' merged blocks...")
    term_merged_ranges = []
    for merged_range_obj in ws.merged_cells.ranges:
        min_col, min_row, max_col, _ = merged_range_obj.bounds
        if min_col == term_col_idx and max_col == term_col_idx: # It's a merge in the Application Term column
            term_merged_ranges.append(merged_range_obj)
    
    # Sort these term blocks by their starting row
    term_merged_ranges.sort(key=lambda r: r.min_row)

    apply_alternate_fill_to_term_block = True # Start with applying fill to the first term block
    for term_range in term_merged_ranges:
        # Skip if this merged range is for "Grand Total" itself
        top_left_cell_value = ws.cell(row=term_range.min_row, column=term_range.min_col).value
        if str(top_left_cell_value).strip() == "Grand Total":
            continue

        if apply_alternate_fill_to_term_block:
            logger.debug(f"Applying alt fill to term block starting at row {term_range.min_row} for term '{top_left_cell_value}'")
            for row_idx_in_merge in range(term_range.min_row, term_range.max_row + 1):
                ws.cell(row=row_idx_in_merge, column=term_col_idx).fill = fill_term_block_alt
        else:
            logger.debug(f"Skipping alt fill for term block starting at row {term_range.min_row} for term '{top_left_cell_value}' (will be white/default)")
        apply_alternate_fill_to_term_block = not apply_alternate_fill_to_term_block # Toggle for next term block
    logger.debug("Alternating fill for 'Application Term' blocks applied.")

    # --- 2. Merge 'Total' (Program) and 'Grand Total' (Term) label cells ---
    logger.debug("Merging 'Total' and 'Grand Total' label cells...")
    for row_idx in range(2, ws.max_row + 1):
        program_cell_value = str(ws.cell(row=row_idx, column=program_col_idx).value).strip()
        term_cell_value = str(ws.cell(row=row_idx, column=term_col_idx).value).strip()

        if program_cell_value == "Total":
            try:
                ws.merge_cells(start_row=row_idx, end_row=row_idx, start_column=program_col_idx, end_column=degree_col_idx)
                ws.cell(row=row_idx, column=program_col_idx).alignment = align_center_center
                logger.debug(f"Merged 'Total' label in row {row_idx}")
            except Exception as e:
                 logger.warning(f"Could not merge 'Total' label in row {row_idx}: {e}")
        
        if term_cell_value == "Grand Total": # This is the 'Grand Total' row identifier
            try:
                # Merge from "Application Term" (Grand Total label) up to "DEGREE" column
                ws.merge_cells(start_row=row_idx, end_row=row_idx, start_column=term_col_idx, end_column=degree_col_idx)
                ws.cell(row=row_idx, column=term_col_idx).alignment = align_center_center
                logger.debug(f"Merged 'Grand Total' label in row {row_idx}")
            except Exception as e:
                 logger.warning(f"Could not merge 'Grand Total' label in row {row_idx}: {e}")
    logger.debug("'Total' and 'Grand Total' label cell merging complete.")

    # --- 3. Apply Styles (Fonts, Fills, Borders, Alignment) ---
    logger.debug("Applying row/cell styles (fonts, fills, borders, alignment)...")
    header_row_idx = 1
    max_col_idx = ws.max_column

    # Style Header Row
    for col_idx in range(1, max_col_idx + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = font_bold
        cell.alignment = align_center_center
        cell.border = Border(
            left=side_medium_black if col_idx == 1 else side_thin_white,
            right=side_medium_black if col_idx == max_col_idx else side_thin_white,
            top=side_medium_black,
            bottom=side_medium_black
        )

    # Style Data Rows (zebra striping, total/grand total highlighting, borders)
    is_even_data_row_group = True # For zebra striping, toggles per non-total data row
    
    for row_idx in range(2, ws.max_row + 1):
        current_row_fill = None
        current_row_font = Font() # Default font
        is_total_row_type = str(ws.cell(row=row_idx, column=program_col_idx).value).strip() == "Total"
        is_grand_total_row_type = str(ws.cell(row=row_idx, column=term_col_idx).value).strip() == "Grand Total"

        if is_grand_total_row_type:
            current_row_fill = fill_grand_total_row
            current_row_font = font_bold
        elif is_total_row_type:
            current_row_fill = fill_total_row
            current_row_font = font_bold
        else: # It's a regular data row
            if is_even_data_row_group:
                 # Apply alt fill only to data columns (e.g., from Program onwards or all if desired)
                 # This example applies to all columns of the data row for simplicity.
                 current_row_fill = fill_alt_row 
            is_even_data_row_group = not is_even_data_row_group # Toggle for next data row

        for col_idx in range(1, max_col_idx + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if current_row_fill:
                # For total/grand_total rows, fill all columns.
                # For data rows, fill all columns if current_row_fill is set (zebra).
                cell.fill = current_row_fill
            
            if current_row_font.bold: # Apply bold font if set (for total/grand_total)
                cell.font = current_row_font

            # Center align numeric/data columns (typically from first_data_col_idx onwards)
            if col_idx >= first_data_col_idx:
                cell.alignment = align_center_center
            
            # Default inner borders to thin white for a cleaner look
            border_left = side_thin_white
            border_right = side_thin_white
            border_top = side_thin_white
            border_bottom = side_thin_white

            # Outer borders for the table
            if col_idx == 1: border_left = side_medium_black
            if col_idx == max_col_idx: border_right = side_medium_black
            # Header row already has top/bottom black. For data rows:
            # if row_idx == 2: border_top = side_medium_black # Top of data area
            if row_idx == ws.max_row: border_bottom = side_medium_black # Bottom of table

            # Special borders for total and grand total rows
            if is_total_row_type or is_grand_total_row_type:
                border_top = side_medium_black
                border_bottom = side_medium_black
            
            cell.border = Border(left=border_left, right=border_right, top=border_top, bottom=border_bottom)
    logger.debug("Row/cell styles applied.")

    # --- 4. Refine borders for merged "Application Term" cells (make their own box) ---
    logger.debug("Refining borders for merged 'Application Term' cells...")
    for merged_range_obj in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row_range = merged_range_obj.bounds
        # Check if this merge is for the 'Application Term' column
        if min_col == term_col_idx and max_col == term_col_idx:
            for r_idx in range(min_row, max_row_range + 1):
                for c_idx in range(min_col, max_col +1): # Should be only one col
                    cell = ws.cell(row=r_idx, column=c_idx)
                    current_border = cell.border.copy() # Get existing border
                    current_border.left = side_medium_black
                    current_border.right = side_medium_black
                    if r_idx == min_row: current_border.top = side_medium_black
                    if r_idx == max_row_range: current_border.bottom = side_medium_black
                    cell.border = current_border
    logger.debug("Borders for merged 'Application Term' cells refined.")

    # --- 5. Auto-adjust column widths as the last step ---
    _auto_adjust_column_widths(ws)
    logger.info(f"Detailed styling for '{ws.title}' complete.")


def format_excel_workbook(excel_path: str):
    """
    Loads an existing Excel workbook and applies formatting to its sheets.
    Specific formatting is applied to the 'Progress Report' sheet.
    Generic formatting (auto-width, basic header) is applied to other sheets.
    """
    logger.info(f"Starting Excel formatting for workbook: {excel_path}")
    try:
        wb = load_workbook(excel_path)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"Formatting sheet: {sheet_name}")

            if sheet_name == settings.EXCEL_SHEET_NAME_PROGRESS_REPORT:
                _apply_styles_to_progress_report_sheet(ws)
            elif sheet_name.startswith("ERROR_"): # Basic formatting for error sheets
                 _auto_adjust_column_widths(ws)
                 if ws.max_row >=1:
                    for cell in ws[1]: # Header
                        cell.font = Font(bold=True, color="FF0000") # Red bold for error header
            else: # Generic formatting for other sheets (e.g., Raw Data)
                logger.info(f"Applying generic formatting to sheet: {sheet_name}")
                # --- Freeze top row for other sheets as well ---
                ws.freeze_panes = 'A2' 
                logger.debug(f"Froze top row for sheet: {ws.title}")

                if ws.max_row >= 1: # Style header if it exists
                    for cell in ws[1]: # Header row
                        cell.font = Font(bold=settings.EXCEL_FONT_BOLD)
                        cell.alignment = Alignment(**settings.EXCEL_ALIGNMENT_CENTER)
                _auto_adjust_column_widths(ws)
        
        wb.save(excel_path)
        logger.info(f"✅ Excel workbook formatting complete. Saved to: {excel_path}")

    except FileNotFoundError:
        logger.error(f"Error: Excel file not found at '{excel_path}'. Cannot apply formatting.")
        raise # Or handle as per application requirements
    except Exception as e:
        logger.error(f"An unexpected error occurred during Excel formatting of '{excel_path}': {e}", exc_info=True)
        raise # Or handle
